import xlwt
import openpyxl
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.checks import messages
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404

from Benefits.forms import BenefitsForm
from Benefits.models import Benefit, Purchase, Wish
from Worker.forms import LoginForm, ProfileForm, UserForm, ScoreForm
from django.urls import reverse

from Worker.models import Profile


def beginning_page(request):
    return render(request, 'index.html')


def home(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect(reverse('personal_cabinet'))
            else:
                form.add_error(None, 'Incorrect username or password')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})


@login_required
def personal_cabinet(request):
    user = request.user
    bought_benefits = Purchase.objects.filter(user=user)
    if user.profile.position == 'hr' or user.profile.position == 'HR':
        return render(request, 'logged_hr.html', {'benefits': bought_benefits})
    else:
        return render(request, 'logged_worker.html', {'benefits': bought_benefits})
    # bought_benefits = Purchase.objects.filter(user=user)
    # param = {'benefits': benefits, 'bought_benefits': bought_benefits, 'wished_benefits': wished_benefits}


@login_required
def wished_show(request):
    user = request.user
    wished_benefits = Wish.objects.filter(user=user)
    return render(request, 'wished_benefits.html', {'benefits': wished_benefits})


@login_required
def benefits_show(request):
    benefits = Benefit.objects.all()
    return render(request, 'benefits.html', {'benefits': benefits})


@login_required
def bought(request, pk):
    user_id = request.user
    benefit = get_object_or_404(Benefit, pk=pk)
    if request.method == 'POST':
        document = request.FILES.get('document')
        if document.content_type != 'application/msword':
            return redirect('benefits_show')
        if user_id.profile.balance >= benefit.price and document:
            user_id.profile.balance -= benefit.price
            new_obj = Purchase(user=user_id, benefit=benefit)
            if Wish.objects.filter(user=user_id, benefit=benefit).exists():
                Wish.objects.filter(user=user_id, benefit=benefit).delete()
            new_obj.save()
            user_id.save()
        else:
            return redirect('benefits_show')
    return redirect('personal_cabinet')


@login_required
def wished(request, pk):
    user_id = request.user
    benefit = get_object_or_404(Benefit, pk=pk)
    if not Wish.objects.filter(user=user_id, benefit=benefit).exists():
        new_obj = Wish(user=user_id, benefit=benefit)
        new_obj.save()
    return redirect('wished_show')


@login_required
def wished_remove(request, pk):
    user_id = request.user
    benefit = get_object_or_404(Benefit, pk=pk)
    Wish.objects.filter(user=user_id, benefit=benefit).delete()
    return redirect('wished_show')


@login_required
def employees_show(request):
    employees = User.objects.all()
    return render(request, 'employees.html', {'employees': employees})


@login_required
def delete_benefit(request, pk):
    Benefit.objects.filter(pk=pk).delete()
    return redirect('benefits_show')


@login_required()
def delete_worker(request, pk):
    User.objects.filter(pk=pk).delete()
    return redirect('employees_show')


@login_required
def sort_by_сity(request):
    benefits = Benefit.objects.all().order_by('city', 'price')
    return render(request, 'benefits.html', {'benefits': benefits})


@login_required
def sort_by_popularity(request):
    benefits = Benefit.objects.annotate(num_purchases=Count('purchases')).order_by('-num_purchases')
    return render(request, 'benefits.html', {'benefits': benefits})


@login_required
def sort_by_price_asc(request):
    benefits = Benefit.objects.all().order_by('price')
    return render(request, 'benefits.html', {'benefits': benefits})


@login_required
def logout(request):
    return redirect('home')


@login_required
def add_employee(request):
    if request.method == 'POST':
        user_form = UserForm(request.POST)
        profile_form = ProfileForm(request.POST)
        if profile_form.is_valid() and user_form.is_valid():
            user = user_form.save()
            user.username = user_form.cleaned_data['username']
            user.set_password(user_form.cleaned_data['password'])
            user.save()
            user.profile.surname = profile_form.cleaned_data['surname']
            user.profile.first_name = profile_form.cleaned_data['first_name']
            user.profile.last_name = profile_form.cleaned_data['last_name']
            user.profile.position = profile_form.cleaned_data['position']
            user.profile.experience = profile_form.cleaned_data['experience']
            user.profile.city = profile_form.cleaned_data['city']
            user.profile.balance = profile_form.cleaned_data['balance']
            user.profile.email = profile_form.cleaned_data['email']
            user.save()
            return redirect('employees_show')
    else:
        profile_form = ProfileForm()
        user_form = UserForm()
    return render(request, 'add_employee.html', {'profile_form': profile_form, 'user_form': user_form})


@login_required
def add_score(request, pk):
    profile = Profile.objects.get(pk=pk)
    user = User.objects.get(pk=pk)
    if request.method == 'POST':
        form = ScoreForm(request.POST)
        if form.is_valid():
            score_to_add = form.cleaned_data['score_to_add']
            profile.balance += score_to_add
            profile.save()
            return redirect('employees_show')
    else:
        form = ScoreForm()
    return render(request, 'add_bonuses.html', {'form': form, 'user': user})


@login_required
def my_view(request, pk):
    if request.method == 'POST':
        document = request.FILES['document']
    return redirect('personal_cabinet')


@login_required
def add_benefit(request):
    if request.method == 'POST':
        form = BenefitsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('benefits_show')
    else:
        form = BenefitsForm()
    return render(request, 'add_benefit.html', {'form': form})



@login_required
def export_users_xls(request):
    purchases_by_user = Purchase.objects.all().select_related('User', 'Benefit').order_by('user__username').values('user__username', 'benefit__name', 'date')
    workbook = openpyxl.Workbook()

    worksheet = workbook.active
    worksheet.title = 'Purchases'

    worksheet['A1'] = 'Сотрудник'
    worksheet['B1'] = 'Льгота'
    worksheet['C1'] = 'Дата'

    row_num = 2
    for purchase in purchases_by_user:
        worksheet.cell(row=row_num, column=1, value=purchase['user__username'])
        worksheet.cell(row=row_num, column=2, value=purchase['benefit__name'])
        worksheet.cell(row=row_num, column=3, value=purchase['date'].strftime('%Y-%m-%d %H:%M:%S'))
        row_num += 1

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=purchases.xlsx'
    workbook.save(response)
    return response


@login_required
def faq(request):
    return render(request, 'faq.html')
